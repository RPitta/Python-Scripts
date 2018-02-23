#! python3
# xlsx2txt - A program that writes the cells of each column into a text file

import openpyxl
from openpyxl.cell import get_column_letter

wb = openpyxl.load_workbook('txt2xlsx.xlsx')
sheet = wb.active

for col in range(1, sheet.max_column + 1):
	txtFile = open('col' + get_column_letter(col) + '.txt', 'a')
	for row in range(1, sheet.max_row + 1):
		txtFile.write(str(sheet.cell(row=row, column=col).value))
	txtFile.close()

