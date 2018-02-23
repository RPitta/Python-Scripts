#! python3
# rowInserter.py - takes two integers and a filename as command line arguments
# 				   and inserts rows based on the integer given.

import openpyxl, os, sys

os.chdir('C:\\Users\\Rodrigo')
oldWb = openpyxl.load_workbook(sys.argv[3])
newWb = openpyxl.Workbook()
oldSheet = oldWb.active
newSheet = newWb.active

for row in range(1, oldSheet.max_row + 1):
	for column in range(1, oldSheet.max_column + 1):
		if row < int(sys.argv[1]):
			newSheet.cell(row=row, column=column).value = oldSheet.cell(row=row, column=column).value
		else:
			newSheet.cell(row=row+int(sys.argv[2]), column=column).value = oldSheet.cell(row=row, column=column).value
			
fileName, fileExtension = os.path.splitext(sys.argv[3])
newWb.save('{0}_with_blanks{1}'.format(fileName, fileExtension))
oldWb.save(sys.argv[3])
