#! python3
# excel2csv.py - Converts excel files to csv files.

import openpyxl, os, csv
os.chdir('C:\\Users\\Rodrigo\\PythonScripts\\xlsx2csv')
for excelFile in os.listdir('.'):
	if excelFile.endswith('.xlsx'):
		wb = openpyxl.load_workbook(excelFile)
		for sheetName in wb.get_sheet_names():
			# Loop through every sheet in the workbook
			sheet = wb.get_sheet_by_name(sheetName)

			# Create the CSV filename from the Excel filename and sheet title.
			xlsxFile, ext,  = os.path.splitext(excelFile) 
			outputFile = open(xlsxFile + '_' + sheetName + '.csv', 'w', newline='')
			# Create the csv.writer object for this CSV file.
			outputWriter = csv.writer(outputFile)
			# Loop through every row in the sheet.
			for rowNum in range(1, sheet.get_highest_row() + 1):
				rowData = []
				# Loop through each cell in the row.
				for colNum in range(1, sheet.get_highest_column() + 1):
					# Append each cell's data to rowData.
					if sheet.cell(row=rowNum, column=colNum).value != None:
						rowData.append(sheet.cell(row=rowNum, column=colNum).value)
					else:
						continue

				print(rowData)

				# Write the rowData list to the CSV file.
				outputWriter.writerow(rowData)

			outputFile.close()