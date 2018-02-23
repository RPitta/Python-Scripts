#! python3
# cell_inverter.py - inverts row and column of cells in a spreadsheet

import openpyxl, sys, os

os.chdir('C:\\Users\\Rodrigo')
wb = openpyxl.load_workbook('example.xlsx')
newWb = openpyxl.Workbook()
sheet = wb.active
newSheet = newWb.active

print(sheet['A1'].value)

sheetData = []

for colNum in range(1, sheet.max_column + 1):
	sheetData.append([])
	for rowNum in range(1, sheet.max_row + 1):
		sheetData[colNum - 1].append(rowNum)

for x in range(1, len(sheetData) + 1):
	for y in range(1, len(sheetData[0]) + 1):
		newSheet.cell(row=x, column=y).value = sheet.cell(row=y, column=x).value
		

newWb.save('example_copy.xlsx')